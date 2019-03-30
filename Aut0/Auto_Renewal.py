from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import openpyxl
import os
import pandas as pd
import numpy as np
import datetime
import time
import glob
import glob
os.chdir(r'C:\Users\esaugph\Desktop\AutoRenewal')
date = datetime.datetime.now()
date1 = date - datetime.timedelta(days = 0)
now = time.time()
old = now-60*60
old_files=glob.glob('renewal*')
for i in old_files:
	stat = os.stat(i)
	if stat.st_ctime < old:
		os.remove(i)
File_Pre=glob.glob('renewal')
os.chdir(r'C:\Users\esaugph\Desktop\AutoRenewal')
North='renewalResponse_Server2_' + str(date1.strftime("%d%m%Y")) + '.csv'
South='renewalResponse_Server1_' + str(date1.strftime("%d%m%Y")) + '.csv'
print North,South
wb = openpyxl.Workbook()
ws = wb.get_sheet_by_name('Sheet')
ws.title = 'Auto_Renewal'
border_thin = Border(left=Side(border_style='thin',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),)
border_medium = Border(left=Side(border_style='medium',
                           color='FF000000'),
                 right=Side(border_style='medium',
                            color='FF000000'),
                 top=Side(border_style='medium',
                          color='FF000000'),
                 bottom=Side(border_style='medium',
                             color='FF000000'),)
font1 = Font(name='Calibri',
             size=11,
             bold=True,
             italic=False,
             vertAlign=None,
             underline='none',
             strike=False,
             color='FF000000')
font2 = Font(name='Calibri',
             size=11,
             bold=False,
             italic=False,
             vertAlign=None,
             underline='none',
             strike=False,
             color='FF000000')
def fill(start,end,i,j,k,l,sheet):
    pinkfill = PatternFill(start_color=start,end_color=end,fill_type='solid')
    a=i
    b=j
    while(a<=k):
        while(b<=l):
            sheet.cell(row=a,column=b).fill=pinkfill
            b=b+1
        b=j
        a=a+1
def border(i,j,k,l,bord,sheet,H,V):
    a=i
    b=j
    while(a<=k):
        while(b<=l):
            sheet.cell(row=a,column=b).border=bord
            sheet.cell(row=a,column=b).alignment=Alignment(horizontal=H,vertical=V)
            b=b+1
        b=j
        a=a+1
def Summary(Hub,r):
    global df
    df_North = pd.read_csv(Hub)
    df=pd.pivot_table(df_North,index=[u" Result Description"],columns=[u" CircleCode"],values=[u" MSISDN"],
               aggfunc='count',fill_value=0)
    df[u' Total']=df.sum(axis=1)
    df.loc[u' Grand Total']=df.sum()
    fileOut = 'Auto_Renewal.xlsx'
    ws.cell(row=1, column=1).value = 'Error Code'
    array=np.array(df)
    cols=df.columns
    index=df.index
    k=1
    x=2
    global y
    for i in range(len(index)+1):
        for k in range(len(cols)+1):
            if(i==0):
                ws.cell(row=r, column=k+1).value=cols[k-1][1]
                ws.cell(row=r, column=k+1).font = font1
            if(i>0 and k==0):
                ws.cell(row=i+r, column=k+1).value=index[i-1]
                ws.cell(row=i+r, column=k+1).font = font1
            if(i>0 and k>0):
                ws.cell(row=i+r, column=k+1).value=array[i-1][k-1]
                ws.cell(row=i+r, column=k+1).font = font2
    ws.cell(row=r, column=1).value = 'Error Code'
    ws.cell(row=r, column=k+1).value = 'Total'
    ws.cell(row=r, column=1).font = font1
    ws.cell(row=r, column=k + 1).font = font1
    border(r,1,i+r,k+1, border_medium, ws,'left','center')
    fill('EEEE00','FFFFFF',r,1,r,k+1, ws)
    fill('9ACD32', 'FFFFFF',i+r,1,i+r,k+1, ws)
    y=i+r+2
Summary(North,1)
Summary(South,y)
wb.save('test.xlsx')




